"""
Existe una base de datos histórica con registros desde enero de 2013:

https://m.inei.gob.pe/estadisticas/indice-tematico/price-indexes/

El link para "Índices Unificados de Precios de la Construcción (Base Julio 1992 = 100)"
no es estático, por lo que hay que descargarlo manualmente. Sin embargo, sirve para
obtener la base de datos con la que empezar.

Se actualiza mensualmente según lo publicado en la página del INEI:

https://www.gob.pe/institucion/inei/informes-publicaciones/4025211-indice-unificados-de-precios-de-la-construccion-para-las-seis-areas-geograficas

"""

import sqlite3
from openpyxl import Workbook, load_workbook


class SQLiteError(Exception):
    pass


def insert_regiones(path: str, regiones: list[tuple]):
    conn = sqlite3.connect(path)

    try:
        conn.setconfig(1002)
        cursor = conn.cursor()
        sql = """
            INSERT INTO regiones("región", "área_geográfica")
            VALUES (?, ?)
        """
        cursor.executemany(sql, regiones)
        conn.commit()
    except sqlite3.Error as e:
        raise SQLiteError(e)
    finally:
        conn.close()


def insert_códigos(path: str, códigos: list[tuple]):
    conn = sqlite3.connect(path)

    try:
        conn.setconfig(1002)
        cursor = conn.cursor()
        sql = """
            INSERT INTO códigos("ID", "Nombre", "Activo")
            VALUES (?, ?, ?)
        """
        cursor.executemany(sql, códigos)
        conn.commit()
    except sqlite3.Error as e:
        raise SQLiteError(e)
    finally:
        conn.close()


def insert_índices(path: str, índices: list[tuple]):
    conn = sqlite3.connect(path)

    try:
        conn.setconfig(1002)
        cursor = conn.cursor()
        sql = """
            INSERT INTO índices_unificados("Índice", "Código", "Área", "Año", "Mes")
            VALUES (?, ?, ?, ?, ?)
    """
        cursor.executemany(sql, índices)
        conn.commit()
    except sqlite3.Error as e:
        raise SQLiteError(e)
    finally:
        conn.close()


def get_regiones(path: str):
    regiones_por_area = []

    with open(path, 'r+') as file:
        lines = file.readlines()
        for line in lines:
            área = line.split(sep=':')[0].strip()[-1]
            regiones = line.split(sep=':')[1].split(sep=',')
            for región in regiones:
                row = región.strip(), int(área)
                regiones_por_area.append(row)
    return regiones_por_area


def get_códigos(path: str):
    categorías = []

    with open(path, 'r+') as file:
        lines = file.readlines()
        for line in lines:
            código = int(line.split(sep=',')[0].strip())
            elemento = line.split(sep=',')[1].strip()
            activo = int(line.split(sep=',')[2].strip())
            row = código, elemento, activo
            categorías.append(row)
    return categorías


def get_precios(path: str):
    meses = {
        'Ene': 1,
        'Feb': 2,
        'Mar': 3,
        'Abr': 4,
        'May': 5,
        'Jun': 6,
        'Jul': 7,
        'Ago': 8,
        'Set': 9,
        'Oct': 10,
        'Nov': 11,
        'Dic': 12
    }
    wb = load_workbook(path)
    sheets = wb.sheetnames
    índices = []
    for sheet in sheets:
        if '_' in sheet or '-' in sheet:
            if '_' in sheet:
                delimiter = '_'
            elif '-' in sheet:
                delimiter = '-'
            mes = sheet.split(sep=delimiter)[0].strip()
            mes = meses[mes]
            año = int(sheet.split(sep=delimiter)[1].strip())
            print(f'Periodo {año}{str(mes).zfill(2)}:')
            ws = wb[sheet]
            índices_left = map_índices(mes, año, ws, 2, 7, 8, 39)
            índices_right = map_índices(mes, año, ws, 9, 14, 8, 43)
            índices = índices + índices_left + índices_right
    return índices


def map_índices(mes: int, año: int, ws, start_col: int, end_col: int, start_row: int, end_row: int):
    índices = []
    for i, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True), start=start_row):
        for j, cell in enumerate(row, start=start_col):
            if type(cell) == str and '*' in cell:
                cell = None
            elif type(cell) == str and ',' in cell:
                cell = cell.replace(',','.')
                cell = float(cell)
            else:
                cell = float(cell)
            line = cell, int(ws.cell(i, start_col - 1).value), int(ws.cell(7, j).value), año, mes
            print(line)
            índices.append(line)

    return índices


def main():
    # regiones = get_regiones(path='índices_unificados/áreas_geográficas.txt')
    # insert_regiones(path='índices_unificados/indices_unificados.db', regiones=regiones)

    # relación_índices = get_códigos('índices_unificados/códigos.txt')
    # insert_códigos('índices_unificados/indices_unificados.db', relación_índices)

    índices = get_precios('índices_unificados/set24.xlsx')
    insert_índices('índices_unificados/índices_unificados.db', índices)
    pass


main()
